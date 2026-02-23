"""
FinSight AI – FastAPI Main Entry Point
All API routes: Auth, Ingestion, Analytics, AI Engine, Assistant, Audit
"""
import os
import sys
import uuid
import logging
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, File, UploadFile, Depends, HTTPException, Request, Form, Body
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, EmailStr
from dotenv import load_dotenv
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

sys.path.insert(0, os.path.dirname(__file__))
load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s — %(message)s")
logger = logging.getLogger(__name__)

ENVIRONMENT = os.getenv("ENVIRONMENT", "development")
_TESTING    = os.getenv("TESTING", "0") == "1"
# In test mode use a very high limit so no test hits 429
_LOGIN_LIMIT   = "10000/minute" if _TESTING else "10/minute"
_REFRESH_LIMIT = "10000/minute" if _TESTING else "20/minute"

from database import DB, init_db
from auth import (
    authenticate_user, create_access_token, create_refresh_token,
    decode_refresh_token, get_current_user, require_permission,
    validate_password_strength
)
from audit.logger import log_action
from constants import ACTION_LOGIN, ACTION_LOGOUT, ACTION_LOGIN_FAILED, ACTION_REGISTER
from ingestion.parser import parse_csv, parse_excel, normalize_dataframe
from ingestion.cleaner import clean
from ingestion.categorizer import categorize
from analytics.cashflow import (
    get_monthly_cashflow, get_daily_cashflow,
    get_category_breakdown, get_income_expense_trends
)
from analytics.health_score import compute_health_score
from ai_engine.forecaster import forecast_cashflow
from ai_engine.anomaly import detect_anomalies
from ai_engine.risk_scorer import compute_risk_score
from ai_engine.recommender import generate_recommendations
from assistant.chat import chat as assistant_chat
from exchange_rates import get_all_rates, convert as fx_convert, get_symbol


# ── Rate Limiter ───────────────────────────────────────────────────────────────
limiter = Limiter(key_func=get_remote_address)

# ── App Setup ─────────────────────────────────────────────────────────────────
app = FastAPI(
    title="FinSight AI",
    description="Intelligent Financial Decision Engine – Analytics, Forecasting & Risk Scoring",
    version="1.2.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
)

app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS — strict origins in production
_ORIGINS = (
    ["http://localhost:5173", "http://127.0.0.1:5173"]
    if ENVIRONMENT == "development"
    else [os.getenv("FRONTEND_URL", "https://finsight.yourdomain.com")]
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def on_startup():
    init_db()


# ── Pydantic Models ────────────────────────────────────────────────────────────
class ChatRequest(BaseModel):
    query: str

class RegisterRequest(BaseModel):
    email: EmailStr
    password: str
    role: str = "analyst"   # admin | analyst | viewer

class RefreshRequest(BaseModel):
    refresh_token: str


class TokenResponse(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"
    role: str


# ── Auth Routes ────────────────────────────────────────────────────────────────
@app.post("/api/auth/login", response_model=TokenResponse, tags=["Auth"])
@limiter.limit(_LOGIN_LIMIT)
async def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends()):
    user = authenticate_user(form_data.username, form_data.password)
    if not user:
        log_action("unknown", ACTION_LOGIN_FAILED, "auth", form_data.username[:64], request.client.host, "failure")
        raise HTTPException(status_code=401, detail="Invalid email or password")

    access  = create_access_token({"sub": user["id"]})
    refresh = create_refresh_token({"sub": user["id"]})
    log_action(user["id"], ACTION_LOGIN, "auth", user["email"][:64], request.client.host)

    return {"access_token": access, "refresh_token": refresh, "token_type": "bearer", "role": user["role"]}


@app.get("/api/auth/me", tags=["Auth"])
async def get_me(current_user: dict = Depends(get_current_user)):
    return {"id": current_user["id"], "email": current_user["email"], "role": current_user["role"]}


@app.post("/api/auth/register", tags=["Auth"])
async def register(
    request: Request,
    body: RegisterRequest,
    current_user: dict = Depends(require_permission("manage_users"))
):
    """Create a new user — Admin only."""
    from auth import get_password_hash

    # Validate password complexity
    try:
        validate_password_strength(body.password)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    existing = DB.fetch_one("SELECT id FROM users WHERE email = ?", (str(body.email).lower().strip(),))
    if existing:
        raise HTTPException(status_code=409, detail="A user with this email already exists")
    if body.role not in ("admin", "analyst", "viewer"):
        raise HTTPException(status_code=422, detail="Role must be admin, analyst, or viewer")

    new_id = str(uuid.uuid4())
    DB.execute(
        "INSERT INTO users (id, email, password_hash, role) VALUES (?,?,?,?)",
        (new_id, str(body.email).lower().strip(), get_password_hash(body.password), body.role)
    )
    log_action(current_user["id"], ACTION_REGISTER, "auth", str(body.email)[:64], request.client.host)
    return {"success": True, "id": new_id, "email": str(body.email), "role": body.role}


@app.post("/api/auth/refresh", response_model=TokenResponse, tags=["Auth"])
@limiter.limit(_REFRESH_LIMIT)
async def refresh_token(request: Request, body: RefreshRequest):
    """Exchange a valid refresh token for a new access token."""
    payload = decode_refresh_token(body.refresh_token)
    user_id = payload.get("sub")
    user = DB.fetch_one("SELECT * FROM users WHERE id = ? AND is_active = 1", (user_id,))
    if not user:
        raise HTTPException(status_code=401, detail="User not found or deactivated")
    access  = create_access_token({"sub": user["id"]})
    refresh = create_refresh_token({"sub": user["id"]})
    return {"access_token": access, "refresh_token": refresh, "token_type": "bearer", "role": user["role"]}


@app.post("/api/auth/logout", tags=["Auth"])
async def logout(request: Request, current_user: dict = Depends(get_current_user)):
    """Logout — client should discard the token."""
    log_action(current_user["id"], "LOGOUT", "auth", current_user["email"], request.client.host)
    return {"success": True, "message": "Logged out successfully"}


# ── User Management Routes (Admin only) ───────────────────────────────────────
class UpdateUserRequest(BaseModel):
    role: Optional[str] = None
    is_active: Optional[bool] = None

class ResetPasswordRequest(BaseModel):
    new_password: str


@app.get("/api/users", tags=["Users"])
async def list_users(
    request: Request,
    current_user: dict = Depends(require_permission("manage_users"))
):
    """List all users — Admin only."""
    rows = DB.fetch_all(
        "SELECT id, email, role, is_active, created_at FROM users ORDER BY created_at DESC"
    )
    return {"data": rows, "total": len(rows)}


@app.patch("/api/users/{user_id}", tags=["Users"])
async def update_user(
    user_id: str,
    request: Request,
    body: UpdateUserRequest,
    current_user: dict = Depends(require_permission("manage_users"))
):
    """Update a user's role or active status — Admin only."""
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot modify your own account here")

    existing = DB.fetch_one("SELECT id FROM users WHERE id = ?", (user_id,))
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")

    if body.role is not None:
        if body.role not in ("admin", "analyst", "viewer"):
            raise HTTPException(status_code=422, detail="Role must be admin, analyst, or viewer")
        DB.execute("UPDATE users SET role = ? WHERE id = ?", (body.role, user_id))
        log_action(current_user["id"], "UPDATE_USER_ROLE", f"users/{user_id}", body.role, request.client.host)

    if body.is_active is not None:
        DB.execute("UPDATE users SET is_active = ? WHERE id = ?", (1 if body.is_active else 0, user_id))
        log_action(current_user["id"], "TOGGLE_USER", f"users/{user_id}", str(body.is_active), request.client.host)

    updated = DB.fetch_one("SELECT id, email, role, is_active FROM users WHERE id = ?", (user_id,))
    return {"success": True, "user": updated}


@app.post("/api/users/{user_id}/reset-password", tags=["Users"])
async def reset_password(
    user_id: str,
    request: Request,
    body: ResetPasswordRequest,
    current_user: dict = Depends(require_permission("manage_users"))
):
    """Reset a user's password — Admin only."""
    from auth import get_password_hash
    if len(body.new_password) < 8:
        raise HTTPException(status_code=422, detail="Password must be at least 8 characters")
    DB.execute("UPDATE users SET password_hash = ? WHERE id = ?",
               (get_password_hash(body.new_password), user_id))
    log_action(current_user["id"], "RESET_PASSWORD", f"users/{user_id}", "", request.client.host)
    return {"success": True}


@app.delete("/api/users/{user_id}", tags=["Users"])
async def delete_user(
    user_id: str,
    request: Request,
    current_user: dict = Depends(require_permission("manage_users"))
):
    """Delete a user — Admin only. Cannot delete yourself."""
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    existing = DB.fetch_one("SELECT email FROM users WHERE id = ?", (user_id,))
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    DB.execute("DELETE FROM users WHERE id = ?", (user_id,))
    log_action(current_user["id"], "DELETE_USER", f"users/{user_id}", existing["email"], request.client.host)
    return {"success": True}


# ── Ingestion Routes ───────────────────────────────────────────────────────────
@app.post("/api/ingest/csv", tags=["Ingestion"])
async def ingest_csv(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("write"))
):
    """Upload and process a CSV transaction file."""
    contents = await file.read()
    return await _process_upload(contents, "csv", current_user, request)


@app.post("/api/ingest/excel", tags=["Ingestion"])
async def ingest_excel(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("write"))
):
    """Upload and process an Excel transaction file."""
    contents = await file.read()
    return await _process_upload(contents, "excel", current_user, request)


async def _process_upload(contents: bytes, fmt: str, user: dict, request: Request) -> dict:
    try:
        raw_df = parse_csv(contents) if fmt == "csv" else parse_excel(contents)
        norm_df = normalize_dataframe(raw_df)
        clean_df, report = clean(norm_df)
        cat_df = categorize(clean_df)

        # Persist to DB
        records = []
        for _, row in cat_df.iterrows():
            records.append((
                row["id"],
                str(row["date"].date()),
                float(row["amount"]),
                row["type"],
                row.get("category", "uncategorized"),
                row.get("sub_category", "other"),
                row.get("description", ""),
                row.get("account_id", "default"),
                row.get("counterparty"),
                row.get("currency", "USD"),
                0,  # is_recurring
                float(row.get("confidence", 0.5)),
                row.get("raw_description", "")
            ))

        DB.executemany(
            """INSERT OR IGNORE INTO transactions
               (id, date, amount, type, category, sub_category, description, account_id,
                counterparty, currency, is_recurring, confidence, raw_description)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            records
        )

        log_action(user["id"], "INGEST", f"transactions/{fmt}", f"{report['final_rows']} rows", request.client.host)
        return {"success": True, "report": report}

    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")


# ── Transaction Routes ─────────────────────────────────────────────────────────
@app.get("/api/transactions", tags=["Transactions"])
async def list_transactions(
    page: int = 1, page_size: int = 50,
    category: Optional[str] = None,
    type: Optional[str] = None,
    account_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(require_permission("read"))
):
    filters = []
    params  = []
    if category:   filters.append("category = ?");    params.append(category)
    if type:       filters.append("type = ?");        params.append(type)
    if account_id: filters.append("account_id = ?"); params.append(account_id)
    if date_from:  filters.append("date >= ?");       params.append(date_from)
    if date_to:    filters.append("date <= ?");       params.append(date_to)

    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    offset = (page - 1) * page_size
    
    rows  = DB.fetch_all(f"SELECT * FROM transactions {where} ORDER BY date DESC LIMIT ? OFFSET ?",
                         tuple(params) + (page_size, offset))
    total = DB.fetch_one(f"SELECT COUNT(*) as cnt FROM transactions {where}", tuple(params))

    return {"data": rows, "total": total["cnt"], "page": page, "page_size": page_size}


@app.delete("/api/transactions/{txn_id}", tags=["Transactions"])
async def delete_transaction(
    txn_id: str, request: Request,
    current_user: dict = Depends(require_permission("delete"))
):
    DB.execute("DELETE FROM transactions WHERE id = ?", (txn_id,))
    log_action(current_user["id"], "DELETE", f"transactions/{txn_id}", "", request.client.host)
    return {"success": True}


# ── Analytics Routes ───────────────────────────────────────────────────────────
@app.get("/api/analytics/overview", tags=["Analytics"])
async def analytics_overview(current_user: dict = Depends(require_permission("read"))):
    """Main KPI overview: health score + key metrics."""
    health = compute_health_score()
    risk   = compute_risk_score()
    return {"health": health, "risk": risk}


@app.get("/api/analytics/cashflow", tags=["Analytics"])
async def analytics_cashflow(
    granularity: str = "monthly",
    days: int = 90,
    current_user: dict = Depends(require_permission("read"))
):
    if granularity == "daily":
        return {"data": get_daily_cashflow(days=days), "granularity": "daily"}
    return {"data": get_monthly_cashflow(), "granularity": "monthly"}


@app.get("/api/analytics/categories", tags=["Analytics"])
async def analytics_categories(current_user: dict = Depends(require_permission("read"))):
    return {"data": get_category_breakdown()}


@app.get("/api/analytics/trends", tags=["Analytics"])
async def analytics_trends(current_user: dict = Depends(require_permission("read"))):
    return {"data": get_income_expense_trends()}


@app.get("/api/analytics/health-score", tags=["Analytics"])
async def analytics_health_score(current_user: dict = Depends(require_permission("read"))):
    return compute_health_score()


# ── AI Engine Routes ───────────────────────────────────────────────────────────
@app.get("/api/ai/forecast", tags=["AI Engine"])
async def ai_forecast(
    periods: int = 6,
    current_user: dict = Depends(require_permission("read"))
):
    return forecast_cashflow(periods=periods)


@app.get("/api/ai/anomalies", tags=["AI Engine"])
async def ai_anomalies(current_user: dict = Depends(require_permission("read"))):
    return {"data": detect_anomalies(), "count": len(detect_anomalies())}


@app.get("/api/ai/risk-score", tags=["AI Engine"])
async def ai_risk_score(current_user: dict = Depends(require_permission("read"))):
    return compute_risk_score()


@app.get("/api/ai/recommendations", tags=["AI Engine"])
async def ai_recommendations(current_user: dict = Depends(require_permission("read"))):
    return {"data": generate_recommendations()}


# ── AI Assistant Route ─────────────────────────────────────────────────────────
@app.post("/api/assistant/chat", tags=["Assistant"])
async def chat(body: ChatRequest, current_user: dict = Depends(require_permission("read"))):
    if not body.query.strip():
        raise HTTPException(status_code=400, detail="Query cannot be empty")
    result = assistant_chat(body.query)
    return result


# ── Audit Route ────────────────────────────────────────────────────────────────
@app.get("/api/audit/logs", tags=["Audit"])
async def get_audit_logs(
    page: int = 1, page_size: int = 50,
    current_user: dict = Depends(require_permission("view_audit"))
):
    offset = (page - 1) * page_size
    rows   = DB.fetch_all("SELECT * FROM audit_logs ORDER BY timestamp DESC LIMIT ? OFFSET ?",
                          (page_size, offset))
    total  = DB.fetch_one("SELECT COUNT(*) as cnt FROM audit_logs")
    return {"data": rows, "total": total["cnt"]}


# ── Health Check ───────────────────────────────────────────────────────────────
@app.get("/api/health", tags=["System"])
def health_check():
    return {"status": "ok", "service": "FinSight AI", "version": "1.3.0"}


# ── Currency Routes ─────────────────────────────────────────────────────────────
@app.get("/api/currency/rates", tags=["Currency"])
async def currency_rates(current_user: dict = Depends(require_permission("read"))):
    """Return all supported currencies with exchange rates (USD base)."""
    return {"base": "USD", "rates": get_all_rates()}


# ── Export Routes ───────────────────────────────────────────────────────────────
@app.get("/api/export/excel", tags=["Export"])
async def export_excel(
    currency: str = "USD",
    current_user: dict = Depends(require_permission("read"))
):
    """Download all transactions as an Excel file (.xlsx)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = DB.fetch_all("SELECT date, amount, type, category, sub_category, description, account_id, currency FROM transactions ORDER BY date DESC")
    symbol = get_symbol(currency)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header styling
    headers = ["Date", "Amount", f"Amount ({currency})", "Type", "Category", "Sub Category", "Description", "Account"]
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="38BDF8")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for r, row in enumerate(rows, 2):
        usd_amount = float(row.get("amount", 0))
        converted  = fx_convert(usd_amount, currency)
        ws.cell(r, 1, row.get("date", ""))
        ws.cell(r, 2, usd_amount)
        ws.cell(r, 3, converted)
        ws.cell(r, 4, row.get("type", ""))
        ws.cell(r, 5, row.get("category", ""))
        ws.cell(r, 6, row.get("sub_category", ""))
        ws.cell(r, 7, row.get("description", ""))
        ws.cell(r, 8, row.get("account_id", ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"finsight_transactions_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/export/pdf", tags=["Export"])
async def export_pdf(
    currency: str = "USD",
    current_user: dict = Depends(require_permission("read"))
):
    """Download a one-page financial summary report as PDF."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm

    health = compute_health_score()
    risk   = compute_risk_score()
    cats   = get_category_breakdown()
    symbol = get_symbol(currency)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle("title", fontSize=20, fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#38BDF8"), spaceAfter=4)
    story.append(Paragraph("FinSight AI — Financial Summary Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')} | Currency: {currency}",
                           ParagraphStyle("sub", fontSize=9, textColor=colors.grey, spaceAfter=16)))

    # KPI Table
    metrics = health.get("metrics", {})
    kpi_data = [
        ["Metric", "Value"],
        ["Health Score",       f"{health.get('score', 0)}/100  ({health.get('level', 'Unknown')})"],
        ["Risk Score",         f"{risk.get('score', 0)}/100   ({risk.get('level', 'Unknown')})"],
        ["Monthly Income",     f"{symbol}{fx_convert(metrics.get('avg_monthly_income', 0), currency):,.0f}"],
        ["Monthly Expenses",   f"{symbol}{fx_convert(metrics.get('avg_monthly_expenses', 0), currency):,.0f}"],
        ["Monthly Net",        f"{symbol}{fx_convert(metrics.get('avg_monthly_net', 0), currency):,.0f}"],
        ["Runway",             f"{metrics.get('runway_months', 0):.1f} months"],
        ["Expense Ratio",      f"{metrics.get('expense_ratio_pct', 0):.1f}%"],
    ]
    kpi_table = Table(kpi_data, colWidths=[8*cm, 8*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#38BDF8")),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#0F172A"), colors.HexColor("#1E293B")]),
        ("TEXTCOLOR",  (0, 1), (-1, -1), colors.white),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#334155")),
        ("PADDING",    (0, 0), (-1, -1), 6),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5*cm))

    # Top Categories
    if cats:
        story.append(Paragraph("Top Expense Categories",
                                ParagraphStyle("h2", fontSize=13, fontName="Helvetica-Bold",
                                               textColor=colors.HexColor("#38BDF8"), spaceBefore=10, spaceAfter=6)))
        cat_data = [["Category", f"Amount ({currency})", "% of Total"]]
        for c in cats[:8]:
            cat_data.append([
                c.get("category", "").title(),
                f"{symbol}{fx_convert(c.get('amount', 0), currency):,.0f}",
                f"{c.get('percent', 0):.1f}%"
            ])
        cat_table = Table(cat_data, colWidths=[7*cm, 5*cm, 4*cm])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#38BDF8")),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#0F172A"), colors.HexColor("#1E293B")]),
            ("TEXTCOLOR",  (0, 1), (-1, -1), colors.white),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#334155")),
            ("PADDING",    (0, 0), (-1, -1), 5),
        ]))
        story.append(cat_table)

    # Footer
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("FinSight AI — Intelligent Financial Decision Engine",
                           ParagraphStyle("footer", fontSize=8, textColor=colors.grey)))

    doc.build(story)
    buf.seek(0)
    filename = f"finsight_report_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
